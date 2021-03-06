#!/usr/bin/python3
# coding=utf-8
# Excel有多个sheet

import os

from openpyxl import load_workbook

proDir = os.path.split(os.path.realpath(__file__))[0]
excelPath = os.path.join(proDir, "../resources/TestCase.xlsx")


class OperationExcel:
    def __init__(self):
        self.open_excel = load_workbook(excelPath)  # 打开Excel表格

    def open_excel_sheet(self, sheet_name):
        """设置需要操作的sheet
        :param sheet_name: 表名
        :return:
        """
        return self.open_excel[sheet_name]

    def get_lines(self, sheet_name):
        """获取Excel表格的总行数
        :param sheet_name:
        :return:
        """
        # line = self.open_excel.max_row
        open_sheet = self.open_excel_sheet(sheet_name)
        line = open_sheet.max_row
        return line

    def from_ab_get_data(self, sheet_name, cell, row):
        """通过单元格获取数据，例如：A2
        :param sheet_name: 表名
        :param cell: 所在列A, B, ...
        :param row: 所在行1, 2, ...
        :return: 对应单元格的值
        """
        open_sheet = self.open_excel_sheet(sheet_name)
        value = open_sheet[cell + str(row)].value
        return value

    def from_xy_get_data(self, sheet_name, x, y):
        """通过单元格坐标获取数据，例如：(1, 2)
        :param sheet_name: 表名
        :param x: 横坐标x
        :param y: 纵坐标y
        :return:返回该坐标(x, y)对应的数据
        """
        open_sheet = self.open_excel_sheet(sheet_name)
        value = open_sheet.cell(x, y).value
        return value

    def write_data(self, sheet_name, cell, row, write_value):
        """写入数据
        :param sheet_name: 表名
        :param cell: 所在列A, B, ...
        :param row: 所在行1, 2, ...
        :param write_value: 写入的值
        :return:
        """
        wb = load_workbook(filename=excelPath)
        ws = wb[sheet_name]
        ws[cell + str(row)] = write_value
        wb.save(filename=excelPath)


if __name__ == "__main__":
    s = OperationExcel()
    for i in range(1, 50):
        d = s.from_ab_get_data('cmb_test_case', 'E', i)
        print(d)
